import openpyxl
import os
import numpy as np
from PIL import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def img_to_array(img):
    image = Image.open("Pictures/" + img)
    image_array = np.array(image.getdata())
    width, height = image.size
    array = []
    counter = 0
    for y in range(0, height):
        array.append([])
        for x in range(0, width):
            array[y].append(image_array[counter])
            counter += 1
    return np.array(array)


def rgb_to_hex(array): return '%02x%02x%02x' % (array[0], array[1], array[2])


def get_percent(x, total): return x * 100 / total


def get_size(w, h):
    if w > 400:
        w = 400
    if h > 400:
        h = 400
    w = 400 / w
    h = 5.94 * 400 / h
    return w, h


def array_to_sheet(array, actual, maximum, name):
    wb = openpyxl.Workbook()
    ws = wb.active
    counter = True
    columnWidth, rowHeight = get_size(len(array), len(array[3]))
    for r in range(1, len(array)):
        ws.insert_rows(r)
        for c in range(1, len(array[r])):
            color = rgb_to_hex(array[r][c])
            ws[get_column_letter(c) + str(r)].fill = (
                PatternFill(start_color=color, end_color=color, fill_type='solid'))
            ws[get_column_letter(c) + str(r)].value = " "
            if counter:
                ws.column_dimensions[get_column_letter(c)].width = columnWidth
        ws.row_dimensions[r].height = rowHeight
        counter = False
        print(str(actual) + " of " + str(maximum) + " // " + str(get_percent(r, len(array))) + " %")
    wb.save("Sheets/" + str(name.split(".")[0]) + ".xlsx")


if __name__ == '__main__':
    cont = 1
    for img in os.listdir("Pictures"):
        a = img_to_array(img)
        array_to_sheet(a, cont, len(os.listdir("Pictures")), img)
        cont += 1